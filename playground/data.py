from baseopensdk import BaseClient
from baseopensdk.api.base.v1 import *
import os, json
import math

APP_TOKEN = os.environ["APP_TOKEN"]
PERSONAL_BASE_TOKEN = os.environ["PERSONAL_BASE_TOKEN"]


def process_data(template: str):
    # 1. build a client
    client: BaseClient = (
        BaseClient.builder()
        .app_token(APP_TOKEN)
        .personal_base_token(PERSONAL_BASE_TOKEN)
        .build()
    )

    # 2. obtain talbles
    list_tables_request = ListAppTableRequest.builder().page_size(100).build()

    list_tables_response = client.base.v1.app_table.list(list_tables_request)

    # print("tables: ", [table.name for table in list_tables_response.data.items])

    TABLE_ID = list_tables_response.data.items[12].table_id

    list_field_request = (
        ListAppTableFieldRequest.builder().page_size(100).table_id(TABLE_ID).build()
    )

    list_field_response = client.base.v1.app_table_field.list(list_field_request)
    fields = getattr(list_field_response.data, "items") or []

    # print(list_tables_response.data.items[12].name, "field:", [field.field_name for field in list_field_response.data.items])

    # 3. get Text fields
    text_field_names = [field.field_name for field in fields]

    print("DateTime fields: ", json.dumps(text_field_names, ensure_ascii=False))

    # 4. iterate through all the records
    list_record_request = (
        ListAppTableRecordRequest.builder()
        .field_names(json.dumps(text_field_names, ensure_ascii=False))
        .page_size(100)
        .table_id(TABLE_ID)
        .build()
    )

    list_record_response = client.base.v1.app_table_record.list(list_record_request)

    records = getattr(list_record_response.data, "items") or []

    print(
        list_tables_response.data.items[12].name,
        "records: size->",
        list_record_response.data.total,
        records[-1].fields,
    )

    records_need_update = []

    # for record in records:
    #     record_id, fields = record.record_id, record.fields
    #     new_fields = {}

    #     for key, value in fields.items():
    #         # replace the value
    #         if key in text_field_names:
    #             new_value = value.replace(source, target)
    #             # add field into new_fields
    #             new_fields[key] = new_value if new_value != value else value

    #     if len(new_fields.keys()) > 0:
    #         records_need_update.append({"record_id": record_id, "fields": new_fields})

    print(records_need_update)

    # 5. batch update records
    batch_update_records_request = (
        BatchUpdateAppTableRecordRequest()
        .builder()
        .table_id(TABLE_ID)
        .request_body(
            BatchUpdateAppTableRecordRequestBody.builder()
            .records(records_need_update)
            .build()
        )
        .build()
    )
    batch_update_records_response = client.base.v1.app_table_record.batch_update(
        batch_update_records_request
    )
    print("success!")


from openpyxl import load_workbook, Workbook


def parse_excel(file_path: str, sheet_title_index={}):
    print("parse_excel: ", file_path)
    # Step 1: Load the workbook
    workbook = load_workbook(filename=file_path)

    # Step 2: Iterate over all sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Sheet Name: {sheet_name}")

        if sheet_name in sheet_title_index and sheet_title_index[sheet_name]<0:
            print(f"{sheet_name}: title index {sheet_title_index[sheet_name]}, skipping")
            continue
        
        # Step 3: Read the first row as fields
        fields = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        print("Fields:", fields)

        title_index = sheet_title_index.get(sheet_name, 1)
        # Optional: Read the remaining data
        row_data = []
        for row in sheet.iter_rows(title_index, values_only=True):
            data = dict(zip(fields, row))
            row_data.append(data)

        # 获取第二行的样式
        source_row = title_index+1
        styles_to_copy = [
            sheet.cell(row=source_row, column=col).style for col in range(1, sheet.max_column + 1)
        ]

        for row in range(3, end_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).style = styles_to_copy[col - 1]

        # 将 data 批量更新到既有的 Excel 中
        for data_item in row_data[-5:]:
            print(data_item)
            sheet.append(list(data_item.values()))

    workbook.save(filename="xxx.xlsx")

    # try:
    #     workbook = load_workbook(filename=file_path)
    # except FileNotFoundError:
    #     # If the file does not exist, create a new one
    #     workbook = Workbook()
    # sheet = workbook.active

    # # Step 4: Write data to the first cell
    # sheet["A1"] = "Hello, Excel!"

    # # Step 5: Save the workbook
    # workbook.save(filename=file_path)

    # target_sheet = workbook.active  # or target_workbook['SheetName']

    # Step 3 & 4: Update the target workbook
    # Assuming you want to append data at the end of the target sheet
    # for data_item in data:
    #     if data_item is not None:
    #         target_sheet.append(list(data_item.values()))

    # # Step 5: Save the workbook
    # workbook.save(filename="target_workbook_path")


if __name__ == "__main__":
    # replace all 'abc' to '233333'
    parse_excel(
        os.path.join(os.path.dirname(__file__), "..", "upload", APP_TOKEN + ".xlsx"),
        title_index={"汇总表": 2, "数据透视表": -1},
    )
    # process_data(os.path.join(os.path.dirname(__file__), "upload", APP_TOKEN+".xlsx"))
